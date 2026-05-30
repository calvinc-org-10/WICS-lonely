import re as regex
from typing import Any, Dict, List, Type
from datetime import date, datetime
from dateutil.parser import parse as dateparse

from PySide6.QtCore import QDate, Qt, Slot
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QDateEdit,
    QWidget,
    QBoxLayout, QVBoxLayout, QGridLayout, 
    QScrollArea, 
    QCheckBox, QGroupBox, QLabel, QPushButton, 
    QProgressBar, 
    )
from sqlalchemy.orm import Session, sessionmaker

from calvincTools.utils import (
    cSRFSingleRecordForm, cFileSelectWidget,
    cExcelFile,
    cstdTabWidget,
    )

from sqlalchemy import (
    and_, select, update, delete, insert, 
    literal_column, 
    text, func, 
    )

# from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel

from qtawesome import icon

from calvincTools.utils.forms import cQFormFieldDef
from calvincTools.utils.forms.definitions.cQFormBtnDef import cQFormBtnDef
from calvincTools.utils.forms.definitions.cQFormFieldDef import cQFormFieldDef
from mathematical_expressions_parser.eval import evaluate

from app.database import Repository, get_app_session, get_app_sessionmaker
from app.models import (
    ActualCounts, CountSchedule, MaterialList, tmpMaterialListUpdate,
    SAP_SOHRecs, SAPPlants_org, UploadSAPResults, 
    )


class UpdateMatlListfromSAP(cSRFSingleRecordForm):
    _ORMmodel = tmpMaterialListUpdate
    _formname = "Update Material List from SAP MM60 or ZMSQV001 Spreadsheet"
    _ssnmaker = get_app_sessionmaker()
    _field_defs = [
        # no fields to edit, everything manually handled
    ]


    def __init__(self, *args, **kwargs):    # copy header from parent?
        self.btnChooseFile = cFileSelectWidget(btnText="Choose Spreadsheet File")
        self.btnChooseFile.fileChosen.connect(self.FileChosen)

        self.dict_chkUpdtOption = {}

        self.chkDeleteIfNotinSprsht = QCheckBox("Delete Records Not in Spreadsheet")

        self.wdgtUpdtStatusText = QLabel("")
        self.wdgtUpdtStatusProgBar = QProgressBar()
        
        super().__init__(*args, **kwargs)   
    # __init__

    def defineFields(self) -> List[cQFormFieldDef]:
        # no fields to edit, everything manually handled
        return []

    ###########################################################
    ############## Override cSimpleRecordForm UI build methods
    ###########################################################

    def _build_fields(self):

        layoutFormFixedTop = self._layouts.fixed_top
        mainFormPage = self.FormPage(0)
        assert isinstance(mainFormPage, QGridLayout)

        chooseFileWidget = QWidget()
        chooseFileLayout = QGridLayout(chooseFileWidget)
        lblChooseFileLabel = QLabel("Choose or Drop SAP Material List Spreadsheet File:")
        chooseFileLayout.addWidget(lblChooseFileLabel, 0, 0)
        chooseFileLayout.addWidget(self.btnChooseFile, 1, 0)

        self.dict_chkUpdtOption['Description'] = QCheckBox("Description")
        self.dict_chkUpdtOption['SAPMatlType'] = QCheckBox("SAP Material Type")
        self.dict_chkUpdtOption['SAPMatlGroup'] = QCheckBox("SAP Material Group")
        self.dict_chkUpdtOption['SAPManuf'] = QCheckBox("SAP Manufacturer")
        self.dict_chkUpdtOption['SAPMPN'] = QCheckBox("SAP Manufacturer Part Number")
        self.dict_chkUpdtOption['SAPABC'] = QCheckBox("SAP ABC Designation")
        self.dict_chkUpdtOption['SAPPrice'] = QCheckBox("Price, Price Unit and Currency")
        grpbxUpdtOptions = QGroupBox('Update Existing Material Records for these Fields:')
        layoutUpdtOptions = QVBoxLayout(grpbxUpdtOptions)
        for chk in self.dict_chkUpdtOption.values():
            layoutUpdtOptions.addWidget(chk)


        mainFormPage.addWidget(chooseFileWidget, 0, 0)
        # mainFormPage.addWidget(PhaseWidget, 2, 0)
        mainFormPage.addWidget(grpbxUpdtOptions, 0, 1, 2, 1)
        mainFormPage.addWidget(self.chkDeleteIfNotinSprsht, 2, 1)

        assert layoutFormFixedTop is not None, "layoutFormFixedTop is required"
        layoutFormFixedTop.addWidget(self.wdgtUpdtStatusText, 0, 0)
        layoutFormFixedTop.addWidget(self.wdgtUpdtStatusProgBar, 1, 0)

    # _placeFields

    def _addActionButtons(self, ActionButtons: List[cQFormBtnDef] | None = None) -> None:
    # def _addActionButtons(self, layoutButtons: QBoxLayout | None = None, layoutHorizontal: bool = True, NavActions: list[tuple[str, QIcon]] | None = None, CRUDActions: list[tuple[str, QIcon]] | None = None) -> None:
        btnUploadFile = QPushButton(icon('fa5s.file-upload'), "Upload Data from Spreadsheet")
        btnUploadFile.clicked.connect(self.uploadFile)
        btnClose = QPushButton(icon('mdi.close-octagon'), "Close Form")
        btnClose.clicked.connect(self.closeForm)

        layoutButtons = self._layouts.buttons
        if layoutButtons is not None:
            layoutButtons.addWidget(btnUploadFile, alignment=Qt.AlignmentFlag.AlignLeft)
            layoutButtons.addStretch(1)
            layoutButtons.addWidget(btnClose, alignment=Qt.AlignmentFlag.AlignRight)
        #endif layoutButtons
    # def _handleActionButton(self, action: str) -> None:
    #     return super()._handleActionButton(action)
    # _add/handleActionButtons

    def initialdisplay(self):
        self.showNewRecordFlag()
        self.showUpdateStatus("")


    ###########################################################
    ############## UI update methods
    ###########################################################

    def showUpdateStatus(self, statusText: str, progressValue: int = 0, progressMax: int = -1) -> None:

        self.wdgtUpdtStatusText.setText(statusText)

        if progressMax < 0:
            self.wdgtUpdtStatusProgBar.setVisible(False)
        else:
            self.wdgtUpdtStatusProgBar.setVisible(True)
            self.wdgtUpdtStatusProgBar.setMaximum(progressMax)
            self.wdgtUpdtStatusProgBar.setValue(progressValue)
        # endif progressMax
    # showUpdateStatus


    ###########################################################
    ############## button responders
    ###########################################################

    @Slot()
    def FileChosen(self):
        # enable the Upload button
        pass

    @Slot()
    def uploadFile(self):
        # disable most of the form while processing

        self.proc_MatlListSAPSprsheet_00InitUMLasync_comm()

        # UMLSSName = self.proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet()     # needed in a client-server environment, not for standalone
        UMLSSName = self.btnChooseFile.getFileChosen()
        self.proc_MatlListSAPSprsheet_01ReadSpreadsheet(UMLSSName)
        self.done_MatlListSAPSprsheet_01ReadSpreadsheet()      # this trigger most of the rest of the processing chain

        # present results to user
        childScreen = ShowUpdateMatlListfromSAPForm()
        childScreen.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        childScreen.show()


        self.proc_MatlListSAPSprsheet_99_Cleanup()
        self.closeForm()

    # uploadFile

    @Slot()
    def closeForm(self):
        # TODO: Implement close form logic
        self.close()


    ###########################################################
    ############## "record" status methods
    ###########################################################

    def isNewRecord(self) -> bool:
        return False


    ###########################################################
    ############## other cSimpleRecordForm overrides
    ###########################################################

    #pylint: disable=unused-argument
    def changeInternalVarField(self, wdgt, intVarField: str, wdgt_value: Any) -> None:
        return


    ###########################################################
    ############## Read / process spreadsheet methods
    ###########################################################

    def proc_MatlListSAPSprsheet_00InitUMLasync_comm(self):
        self.showUpdateStatus('Initializing ...')
    # proc_MatlListSAPSprsheet_00InitUMLasync_comm

    # not needed in standalone version
    # def proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(self):
    #     ...

    def proc_MatlListSAPSprsheet_01ReadSpreadsheet(self, fName):
        # tmpMaterialListUpdate.objects.using(dbToUse).all().delete() - from client-server Django version
        Repository(get_app_sessionmaker(), tmpMaterialListUpdate).removewhere(True)

        self.showUpdateStatus('Reading Spreadsheet')

        wb = cExcelFile.load_from_file(fName, read_only=True)
        assert wb is not None, "Failed to load spreadsheet file."
        ws = wb.active
        assert ws is not None, "Failed to load active worksheet from spreadsheet."
        SAPcolmnNames = ws[1]
        SAPcol = {'Plant':None,'Material': None}
        SAP_SSName_TableName_map = {
                'Material': 'Material',
                'Material description': 'Description',
                'Plant': 'Plant', 'Plnt': 'Plant',
                'Material type': 'SAPMaterialType',  'MTyp': 'SAPMaterialType',
                'Material Group': 'SAPMaterialGroup', 'Matl Group': 'SAPMaterialGroup',
                'Manufact.': 'SAPManuf',
                'MPN': 'SAPMPN',
                'ABC': 'SAPABC', 'ABC Indicator': 'SAPABC',
                'Price': 'Price', 'Standard price': 'Price',
                'Price unit': 'PriceUnit', 'per': 'PriceUnit',
                'Currency':'Currency',
                }
        dict_SAPPlants = {rec.SAPPlant: rec.org_id  for rec in Repository(get_app_sessionmaker(), SAPPlants_org).get_all()}  # preload SAPPlants_org cache
        for col in SAPcolmnNames:
            if col.value in SAP_SSName_TableName_map:
                SAPcol[SAP_SSName_TableName_map[col.value]] = col.column - 1 # type: ignore
        if (SAPcol['Material'] == None or SAPcol['Plant'] == None):
            self.showUpdateStatus('SAP Spreadsheet has bad header row. Plant and/or Material is missing.  See Calvin to fix this.', 0, -1)

            wb.close()
            return      # need to do more than this, but for now, just exit
        # endif bad header row

        numrows = ws.max_row
        nRows = 0
        intrval_announce = min(100, int(max(1, numrows // 10)))
        for row in ws.iter_rows(min_row=2, values_only=True):
            nRows += 1
            if nRows % intrval_announce == 0:
                self.showUpdateStatus(f'Reading Spreadsheet ... record {nRows} of {numrows}', nRows, numrows)

            if row[SAPcol['Material']]==None: MatNum = ''
            else: MatNum = row[SAPcol['Material']]
            validTmpRec = False
            ## create a blank tmpMaterialListUpdate record,
            newrec = tmpMaterialListUpdate()
            if regex.match(".*[\n\t\xA0].*",str(MatNum)):
                validTmpRec = True
                ## refuse to work with special chars embedded in the MatNum
                newrec.recStatus = 'err-MatlNum'
                newrec.errmsg = f'error: {MatNum!a} is an unusable part number. It contains invalid characters and cannot be added to WICS'
            elif len(str(MatNum)):
                validTmpRec = True
                newrec.org_id = dict_SAPPlants.get(str(row[SAPcol['Plant']]), 0)
            # endif invalid Material
            if validTmpRec:
                ## populate by looping through SAPcol,
                ## then save
                for dbColName, ssColNum in SAPcol.items():
                    setattr(newrec,dbColName,row[ssColNum]) # type: ignore
                Repository(get_app_sessionmaker(), tmpMaterialListUpdate).add(newrec)
        # endfor

        wb.close()
        return      # need to do more than this, but for now, just exit

    # proc_MatlListSAPSprsheet_01ReadSpreadsheet

    def done_MatlListSAPSprsheet_01ReadSpreadsheet(self):
        # handle fatal err if occurred during reading
        # statecode = async_comm.objects.using(dbToUse).get(pk=reqid).statecode
        # if statecode != 'fatalerr':
        #     set_async_comm_state(
        #         dbToUse,
        #         statecode = 'done-rdng-sprsht',
        #         statetext = f'Finished Reading Spreadsheet',
        #         )

        self.proc_MatlListSAPSprsheet_02_identifyexistingMaterial()
    # done_MatlListSAPSprsheet_01ReadSpreadsheet

    def proc_MatlListSAPSprsheet_02_identifyexistingMaterial(self):
        self.showUpdateStatus('Identifying Existing Materials ...')
        # UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
        # UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id, '
        # UpdMaterialLinkSQL += "     WICS_tmpmateriallistupdate.recStatus = 'FOUND' "
        # UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
        # UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
        # with connections[dbToUse].cursor() as cursor:
        #     cursor.execute(UpdMaterialLinkSQL)
        # this is a little too wild for the Repository pattern, so we do it manually here
        with get_app_session() as session:
            stmt = (
                update(tmpMaterialListUpdate)
                .values(
                    MaterialLink=select(MaterialList.id)
                        .where(
                            tmpMaterialListUpdate.org_id == MaterialList.org_id,
                            tmpMaterialListUpdate.Material == MaterialList.Material
                        )
                        .correlate(tmpMaterialListUpdate)
                        .scalar_subquery(),
                    recStatus='FOUND'
                )
                .where(
                    select(MaterialList.id)
                        .where(
                            tmpMaterialListUpdate.org_id == MaterialList.org_id,
                            tmpMaterialListUpdate.Material == MaterialList.Material
                        )
                        .correlate(tmpMaterialListUpdate)
                        .exists()
                )
            )

            session.execute(stmt)
            session.commit()
        # endwith

        self.showUpdateStatus('Identifying WICS Materials no longer in SAP MM60 Materials')
        # MustKeepMatlsSelCond = ''
        # MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
        # MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT tmucopy.MaterialLink_id AS Material_id FROM WICS_tmpmateriallistupdate tmucopy WHERE tmucopy.MaterialLink_id IS NOT NULL)'
        # MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
        # MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_actualcounts)'
        # MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
        # MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_countschedule)'
        # MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
        # MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_sap_sohrecs)'

        # DeleteMatlsSelectSQL = "INSERT INTO WICS_tmpmateriallistupdate (recStatus, delMaterialLink, MaterialLink_id, org_id, Material, Description, Plant "
        # DeleteMatlsSelectSQL += ", SAPMaterialType, SAPMaterialGroup, Currency  ) "    # these can go once I set null=True on these fields
        # DeleteMatlsSelectSQL += " SELECT  concat('DEL ',FORMAT(id,0)), id, NULL, org_id, Material, Description, Plant "
        # DeleteMatlsSelectSQL += ", SAPMaterialType, SAPMaterialGroup, Currency  "    # these can go once I set null=True on these fields
        # DeleteMatlsSelectSQL += " FROM WICS_materiallist"
        # DeleteMatlsSelectSQL += f" WHERE ({MustKeepMatlsSelCond})"
        with get_app_session() as session:
            # Build the NOT IN subqueries
            not_in_tmp = select(tmpMaterialListUpdate.MaterialLink).where(
                tmpMaterialListUpdate.MaterialLink.is_not(None)
            ).distinct()

            not_in_actual = select(ActualCounts.Material_id).distinct()
            not_in_schedule = select(CountSchedule.Material_id).distinct()
            not_in_sap = select(SAP_SOHRecs.Material_id).distinct()

            # Build the SELECT statement for materials to delete
            select_stmt = select(
                ('DEL ' + func.format(MaterialList.id, 0)).label('recStatus'),
                MaterialList.id.label('delMaterialLink'),
                literal_column('NULL').label('MaterialLink'),
                MaterialList.org_id,
                MaterialList.Material,
                MaterialList.Description,
                MaterialList.Plant,
                MaterialList.SAPMaterialType,
                MaterialList.SAPMaterialGroup,
                MaterialList.Currency
            ).where(
                MaterialList.id.not_in(not_in_tmp),
                MaterialList.id.not_in(not_in_actual),
                MaterialList.id.not_in(not_in_schedule),
                MaterialList.id.not_in(not_in_sap)
            )

            # INSERT ... SELECT
            stmt = insert(tmpMaterialListUpdate).from_select(
                ['recStatus', 'delMaterialLink', 'MaterialLink', 'org_id', 'Material',
                'Description', 'Plant', 'SAPMaterialType', 'SAPMaterialGroup', 'Currency'],
                select_stmt
            )

            session.execute(stmt)
            session.commit()

        self.showUpdateStatus('Identifying SAP MM60 Materials new to WICS')
        # MarkAddMatlsSelectSQL = "UPDATE WICS_tmpmateriallistupdate"
        # MarkAddMatlsSelectSQL += " SET recStatus = 'ADD'"
        # MarkAddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus is NULL)"
        # Repository(get_app_sessionmaker(), tmpMaterialListUpdate).updatewhere(
        #     lambda rec: rec.MaterialLink_id is None and (rec.recStatus is None),
        #     {'recStatus': 'ADD'}
        # )
        Repository(get_app_sessionmaker(), tmpMaterialListUpdate).updatewhere(
            (tmpMaterialListUpdate.MaterialLink == None) & (tmpMaterialListUpdate.recStatus == None),
            {'recStatus': 'ADD'}
        )

        self.done_MatlListSAPSprsheet_02_identifyexistingMaterial()
    # proc_MatlListSAPSprsheet_02_identifyexistingMaterial
    def done_MatlListSAPSprsheet_02_identifyexistingMaterial(self):
        self.showUpdateStatus('Finished Linking SAP MM60 list to existing WICS Materials ...')

        self.proc_MatlListSAPSprsheet_03_UpdateExistingRecs()
    # done_MatlListSAPSprsheet_02_identifyexistingMaterial

    def proc_MatlListSAPSprsheet_03_UpdateExistingRecs(self):
        def setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(fldName):
            self.showUpdateStatus(f'Updating _{fldName}_ Field in Existing Records')
        # setstate_MatlListSAPSprsheet_03_UpdateExistingRecs

        setstate_MatlListSAPSprsheet_03_UpdateExistingRecs('')

        # (Form Name, db fld Name, zero/blank value)
        # NOTE: SAPPrice updates three fields: Price, PriceUnit, Currency - that's why I don't use a simple dict here
        FormTodbFld_map = [
            ('Description','Description','""'),
            ('SAPMatlType','SAPMaterialType','""'),
            ('SAPMatlGroup','SAPMaterialGroup','""'),
            ('SAPManuf','SAPManuf','""'),
            ('SAPMPN','SAPMPN','""'),
            ('SAPABC','SAPABC','""'),
            ('SAPPrice','Price',0),
            ('SAPPrice','PriceUnit',0),
            ('SAPPrice','Currency','""'),
        ]

        UpdateExistFldSet = {
            key for key, chkbox in self.dict_chkUpdtOption.items()
                if chkbox.isChecked()
        }

        if UpdateExistFldSet:
            MatlList_tbl = MaterialList.__tablename__
            tmpMatlListUpd_tbl = tmpMaterialListUpdate.__tablename__
            for formName, dbName, zeroVal in FormTodbFld_map:
                if formName in UpdateExistFldSet:
                    setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(dbName)
                    # UPDATE this field
                    # for SQLite, the SQLAlchemy way is complex, so we do it manually here
                    UpdSQLSetStmt = f"{dbName}={tmpMatlListUpd_tbl}.{dbName}"
                    UpdSQLWhereStmt = f"(IFNULL({tmpMatlListUpd_tbl}.{dbName},{zeroVal}) != {zeroVal} AND IFNULL({MatlList_tbl}.{dbName},{zeroVal})!=IFNULL({tmpMatlListUpd_tbl}.{dbName},{zeroVal}))"

                    UpdSQLStmt = f"UPDATE {MatlList_tbl}"
                    UpdSQLStmt += f" SET {UpdSQLSetStmt}"
                    UpdSQLStmt += f" FROM {tmpMatlListUpd_tbl}"
                    UpdSQLStmt += f" WHERE ({tmpMatlListUpd_tbl}.MaterialLink={MatlList_tbl}.id) AND {UpdSQLWhereStmt}"
                    UpdSQLStmt += ";"

                    with get_app_session() as session:
                        session.execute(text(UpdSQLStmt))
                        session.commit()
                #endif formName in UpdateExistFldList
            #endfor
        # endif UpdateExistFldList not empty
        self.done_MatlListSAPSprsheet_03_UpdateExistingRecs()
    # proc_MatlListSAPSprsheet_03_UpdateExistingRecs
    def done_MatlListSAPSprsheet_03_UpdateExistingRecs(self):
        self.showUpdateStatus('Finished Updating Existing Records to MM60 values')

        self.proc_MatlListSAPSprsheet_04_Remove()
    # done_MatlListSAPSprsheet_03_UpdateExistingRecs

    def proc_MatlListSAPSprsheet_04_Remove(self):
        if not self.chkDeleteIfNotinSprsht.isChecked():
            self.done_MatlListSAPSprsheet_04_Remove()
            return

        self.showUpdateStatus('Removing WICS Materials no longer in SAP MM60 Materials')

        # do the Removals
        with get_app_session() as session:
            # SQLAlchemy delete with subquery
            subq = select(tmpMaterialListUpdate.delMaterialLink).where(tmpMaterialListUpdate.recStatus.like('DEL%'))
            stmt = delete(MaterialList).where(MaterialList.id.in_(subq))
            session.execute(stmt)
            session.commit()

        self.done_MatlListSAPSprsheet_04_Remove()
    # proc_MatlListSAPSprsheet_04_Remove
    def done_MatlListSAPSprsheet_04_Remove(self):
        self.showUpdateStatus('Finished Removing WICS Materials no longer in SAP MM60 Materials')

        self.proc_MatlListSAPSprsheet_04_Add()
    # done_MatlListSAPSprsheet_04_Remove

    def proc_MatlListSAPSprsheet_04_Add(self):
        self.showUpdateStatus('Adding New WICS Materials from SAP MM60 Materials')

        # do the Additions
        with get_app_session() as session:
             # Select compatible columns from tmp where recStatus is 'ADD'
            select_stmt = select(
                tmpMaterialListUpdate.org_id,
                tmpMaterialListUpdate.Material,
                tmpMaterialListUpdate.Description,
                tmpMaterialListUpdate.Plant,
                tmpMaterialListUpdate.SAPMaterialType,
                tmpMaterialListUpdate.SAPMaterialGroup,
                tmpMaterialListUpdate.SAPManuf,
                tmpMaterialListUpdate.SAPMPN,
                tmpMaterialListUpdate.SAPABC,
                tmpMaterialListUpdate.Price,
                tmpMaterialListUpdate.PriceUnit,
                tmpMaterialListUpdate.Currency
            ).where(
                tmpMaterialListUpdate.recStatus.like('ADD')
            )

            # Insert into MaterialList
            stmt = insert(MaterialList).from_select(
                ['org_id', 'Material', 'Description', 'Plant',
                 'SAPMaterialType', 'SAPMaterialGroup', 'SAPManuf',
                 'SAPMPN', 'SAPABC', 'Price', 'PriceUnit', 'Currency'],
                select_stmt
            )

            session.execute(stmt)
            session.commit()

        self.done_MatlListSAPSprsheet_04_Add()
    # proc_MatlListSAPSprsheet_04_Add
    def done_MatlListSAPSprsheet_04_Add(self):
        self.showUpdateStatus('Finished Adding WICS Materials from SAP MM60 Materials')

        self.proc_MatlListSAPSprsheet_99_FinalProc()
    # done_MatlListSAPSprsheet_04_Add

    def proc_MatlListSAPSprsheet_99_FinalProc(self):
        self.showUpdateStatus('Finished Processing Spreadsheet')
    # proc_MatlListSAPSprsheet_99_FinalProc

    def proc_MatlListSAPSprsheet_99_Cleanup(self):
        # kill async_comm object
        Repository(get_app_sessionmaker(), tmpMaterialListUpdate).removewhere(True)
    # proc_MatlListSAPSprsheet_99_Cleanup
# UpdateMatlListfromSAP
    def end_of_class(self):
        pass

class ShowUpdateMatlListfromSAPForm(QWidget):
    """A form to show the Update Material List from SAP MM60 or ZMSQV001 Spreadsheet form."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.setWindowTitle("Update Material List from SAP MM60 or ZMSQV001 Spreadsheet - Results")
        myLayout = QVBoxLayout(self)
        lblResultsTitle = QLabel("Update Material List from SAP MM60 or ZMSQV001 Spreadsheet - Results")
        myLayout.addWidget(lblResultsTitle)

        wdgtMainArea = QWidget()
        layoutMainArea = QVBoxLayout(wdgtMainArea)
        wdgtScrollArea = QScrollArea()
        wdgtScrollArea.setWidgetResizable(True)
        wdgtScrollArea.setWidget(wdgtMainArea)
        myLayout.addWidget(wdgtScrollArea)

        listImportErrors = Repository(get_app_sessionmaker(), tmpMaterialListUpdate).get_all(
            tmpMaterialListUpdate.recStatus is not None and tmpMaterialListUpdate.recStatus.startswith('err-')
        )
        kountImportErrors = len(listImportErrors)
        if listImportErrors:
            lblErrorsTitle = QLabel(f"{kountImportErrors} errors were encountered during the update:")
            layoutMainArea.addWidget(lblErrorsTitle)
            for errRec in listImportErrors:
                lblErr = QLabel(f"Material: {errRec.Material}, Error: {errRec.errmsg}")
                layoutMainArea.addWidget(lblErr)
        else:
            lblNoErrors = QLabel("No errors were encountered during the update.")
            layoutMainArea.addWidget(lblNoErrors)
        # endif ImpErrList

        listAdditions = Repository(get_app_sessionmaker(), tmpMaterialListUpdate).get_all(
            tmpMaterialListUpdate.recStatus == 'ADD'
        )
        kountAdditions = len(listAdditions)
        if listAdditions:
            lblAdditionsTitle = QLabel(f"{kountAdditions} materials were added to WICS:")
            layoutMainArea.addWidget(lblAdditionsTitle)
            for addRec in listAdditions:
                lblAdd = QLabel(f"id {addRec.id}, Material: (org {addRec.org_id}) {addRec.Material}, {addRec.Description}")
                layoutMainArea.addWidget(lblAdd)
        else:
            lblNoAdditions = QLabel("No new materials were added to WICS.")
            layoutMainArea.addWidget(lblNoAdditions)
        # endif AddList

        listRemovals = Repository(get_app_sessionmaker(), tmpMaterialListUpdate).get_all(
            tmpMaterialListUpdate.recStatus is not None and tmpMaterialListUpdate.recStatus.startswith('DEL ')
        )
        kountRemovals = len(listRemovals)
        if listRemovals:
            lblRemovalsTitle = QLabel(f"{kountRemovals} materials were removed from WICS:")
            layoutMainArea.addWidget(lblRemovalsTitle)
            for delRec in listRemovals:
                lblDel = QLabel(f"Material: (org {delRec.org_id}) {delRec.Material}, {delRec.Description}")
                layoutMainArea.addWidget(lblDel)
        else:
            lblNoRemovals = QLabel("No materials were removed from WICS.")
            layoutMainArea.addWidget(lblNoRemovals)
        # endif RemovalsList
    # __init__
# ShowUpdateMatlListfromSAPForm
    def end_of_class(self):
        pass

class UploadActCountSprsht(cSRFSingleRecordForm):
    """
    A form for uploading and processing count entry spreadsheets into the ActualCounts database.
    This class provides a user interface for selecting an Excel spreadsheet containing count data,
    validating the data, and importing it into the ActualCounts table. It handles field validation,
    material lookup, organization ID resolution, and error tracking during the upload process.
    The spreadsheet must contain a 'Counts' worksheet with required columns:
    - Material: Material number
    - CountDate: Date of the count
    - Counter: Person who performed the count
    - LOCATION: Location of the count
    - Either LocationOnly or CTD_QTY_Expr must be present
    Additional optional columns include:
    - org_id: Organization ID
    - TypicalContainerQty, TypicalPalletQty: Typical quantities
    - Notes, PKGID_Desc, TAGQTY: Additional information
    - Poss Not Rcvd, Mvmt Dur Ct: Status flags
    The upload process:
    1. User selects a spreadsheet file
    2. File is validated for proper format and required columns
    3. Each row is processed, validating data types and required fields
    4. Records are added to ActualCounts table
    5. Results are displayed showing successful imports and any errors
    Attributes:
        _ORMmodel: The ORM model class (ActualCounts)
        _formname: Display name of the form
        _ssnmaker: Session maker for database operations
        fieldDefs: Empty dict as all fields are manually handled
        btnChooseFile: File selection widget
        wdgtUpdtStatusText: Label showing current status
        wdgtUpdtStatusProgBar: Progress bar for upload operations
    """
    _ORMmodel = ActualCounts
    _formname = "Upload Count Entry Spreadsheet"
    _ssnmaker = get_app_sessionmaker()
    fieldDefs = {
        # no fields to edit, everything manually handled
    }

    # pylint: disable=keyword-arg-before-vararg
    def __init__(self, formname: str | None = None, field_defs: List[cQFormFieldDef] | None = None, model: Type[Any] | None = None, ssnmaker: sessionmaker | None = None, parent: QWidget | None = None, *args, **kwargs):
        self.btnChooseFile = cFileSelectWidget(btnText="Choose Spreadsheet File")
        self.btnChooseFile.fileChosen.connect(self.FileChosen)

        self.wdgtUpdtStatusText = QLabel("")
        self.wdgtUpdtStatusProgBar = QProgressBar()

        super().__init__(formname, field_defs, model, ssnmaker, parent, *args, **kwargs)
    # __init__


    def defineFields(self) -> List[cQFormFieldDef]:
        # no fields to edit, everything manually handled
        return []

    ###########################################################
    ############## Override cSimpleRecordForm UI build methods
    ###########################################################

    def _build_fields(self):
    
    # @Slot()
    # def _placeFields(self, layoutFormPages: QTabWidget, layoutFormFixedTop: QGridLayout | None, layoutFormFixedBottom: QGridLayout | None, lookupsAllowed: bool = True) -> None:
        # no fields to place, everything manually handled
        mainFormPage = self.FormPage(0)
        assert isinstance(mainFormPage, QGridLayout)
        layoutFormFixedTop = self._layouts.fixed_top

        chooseFileWidget = QWidget()
        chooseFileLayout = QGridLayout(chooseFileWidget)
        lblChooseFileLabel = QLabel("Choose or Drop Count Entry Spreadsheet File:")
        chooseFileLayout.addWidget(lblChooseFileLabel, 0, 0)
        chooseFileLayout.addWidget(self.btnChooseFile, 1, 0)

        mainFormPage.addWidget(chooseFileWidget, 0, 0)

        assert layoutFormFixedTop is not None, "layoutFormFixedTop is required"
        layoutFormFixedTop.addWidget(self.wdgtUpdtStatusText, 0, 0)
        layoutFormFixedTop.addWidget(self.wdgtUpdtStatusProgBar, 1, 0)
    # _placeFields

    #pylint: disable=signature-differs
    def _addActionButtons(self, ActionButtons) -> None:
        btnUploadFile = QPushButton(icon('fa5s.file-upload'), "Upload Counts")
        btnUploadFile.clicked.connect(self.uploadFile)
        btnClose = QPushButton(icon('mdi.close-octagon'), "Close Form")
        btnClose.clicked.connect(self.closeForm)

        layoutButtons = self._layouts.buttons
        if layoutButtons is not None:
            layoutButtons.addWidget(btnUploadFile, alignment=Qt.AlignmentFlag.AlignLeft)
            layoutButtons.addStretch(1)
            layoutButtons.addWidget(btnClose, alignment=Qt.AlignmentFlag.AlignRight)
        #endif layoutButtons
    # def _handleActionButton(self, action: str) -> None:
    #     return super()._handleActionButton(action)
    # _add/handleActionButtons

    def initialdisplay(self):
        self.showNewRecordFlag()
        self.showUpdateStatus("")
    # initialdisplay

    ###########################################################
    ############## UI update methods
    ###########################################################

    def showUpdateStatus(self, statusText: str, progressValue: int = 0, progressMax: int = -1) -> None:

        self.wdgtUpdtStatusText.setText(statusText)

        if progressMax < 0:
            self.wdgtUpdtStatusProgBar.setVisible(False)
        else:
            self.wdgtUpdtStatusProgBar.setVisible(True)
            self.wdgtUpdtStatusProgBar.setMaximum(progressMax)
            self.wdgtUpdtStatusProgBar.setValue(progressValue)
        # endif progressMax
    # showUpdateStatus


    ###########################################################
    ############## button responders
    ###########################################################

    @Slot()
    def FileChosen(self):
        # enable the Upload button
        pass

    @Slot()
    def uploadFile(self):
        # disable most of the form while processing

        self.proc_UpActCountSprsheet_00InitUpld()

        USSName = self.proc_UpActCountSprsheet_00CopySpreadsheet()
        self.proc_UpActCountSprsheet_01ReadSheet(USSName)
        self.done_UpActCountSprsheet_01ReadSheet()      # this triggers most of the rest of the processing chain

        # present results to user
        childScreen = ShowUploadedCountEntries()
        childScreen.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        childScreen.show()


        self.proc_UpActCountSprsheet_99_Cleanup()
        self.closeForm()


    @Slot()
    def closeForm(self):
        # TODO: Implement close form logic
        self.close()


    ###########################################################
    ############## "record" status methods
    ###########################################################

    def isNewRecord(self) -> bool:
        return False


    ###########################################################
    ############## other cSimpleRecordForm overrides
    ###########################################################

    #pylint: disable=unused-argument
    def changeInternalVarField(self, wdgt, intVarField: str, wdgt_value: Any) -> None:
        return


    ###########################################################
    ############## Read / process spreadsheet methods
    ###########################################################

    def cleanupfld(self, fld, val, CountSprshtDateEpoch = WINDOWS_EPOCH):
        """
        fld is the name of the field in the ActualCount or MaterialList table
        val is the value to be cleaned for insertion into the fld
        Returns  {'usefld':usefld, 'cleanval': cleanval}
            usefld is a boolean indicating that val could/not be cleaned to the correct type
            cleanval is val in the correct type (if usefld==True)
        """
        cleanval = None

        if   fld == 'CountDate':
            if isinstance(val,(date, datetime)):
                usefld = True
                cleanval = val.date() if isinstance(val,datetime) else val
            elif isinstance(val,int):
                usefld = True
                cleanval = from_excel(val, CountSprshtDateEpoch)
            else:
                try:
                    cleanval = dateparse(val).date()
                    usefld = True
                except ValueError:
                    usefld = False
                #endtry
            #endif val type
        elif fld in \
            ['CTD_QTY_Expr',
                ]:
            if isinstance(val,str):
                if val[0] == '=':
                    val = val[1:]
            try:
                v = evaluate(str(val))
            except (SyntaxError, NameError, TypeError, ZeroDivisionError):
                v = "-- INVALID --"
            usefld = (v!="-- INVALID --")
            cleanval = str(val) if (v != "--INVALID--") else None
        elif fld in \
            ['org_id',
                'LocationOnly',
                'FLAG_PossiblyNotRecieved',
                'FLAG_MovementDuringCount',
                ]:
            try:
                cleanval = int(val)
                usefld = True
            except (ValueError, TypeError):
                usefld = False
        elif fld in \
            ['Material',
                'Counter',
                'LOCATION',
                'Notes',
                'TypicalContainerQty',
                'TypicalPalletQty',
                'PKGID_Desc',
                'TAGQTY',
                ]:
            usefld = (val is not None)
            if usefld: cleanval = str(val)
        else:
            usefld = True
            cleanval = val

        return {'usefld':usefld, 'cleanval': cleanval}
    #end def cleanupfld

    def proc_UpActCountSprsheet_00InitUpld(self) -> None:
        self.showUpdateStatus("Initializing Upload of Count Entry Spreadsheet...")
        Repository(get_app_sessionmaker(), UploadSAPResults).removewhere(True)

    def proc_UpActCountSprsheet_00CopySpreadsheet(self) -> str:
        return self.btnChooseFile.getFileChosen()

    def proc_UpActCountSprsheet_01ReadSheet(self, fName: str) -> None:
        self.showUpdateStatus("Reading Count Entry Spreadsheet...")

        NOTdbFld_flags = ['**NOTdbFld**',]

        wb = cExcelFile.load_from_file(fName, read_only=True)
        assert wb is not None, "Failed to load spreadsheet file"
        CountSprshtDateEpoch = wb.epoch
        if 'Counts' in wb:
            ws = wb['Counts']
        else:
            self.showUpdateStatus("Error: This workbook does not contain a sheet named Counts in the correct format")
            wb.close()
            return
        #endif 'Counts' in wb

        SprshtcolmnNames = ws[1]
        SprshtREQUIREDFLDS = ['Material','CountDate','Counter','LOCATION']
            # LocationOnly/CTD_QTY_Expr handled separately since at least one must be present and both can be
        SprshtcolmnMap = {}
        Sprsht_SSName_TableName_map = {
                'CountDate': 'CountDate',
                'Counter': 'Counter',
                'LOCATION': 'LOCATION',
                'org_id': 'org_id',
                'Material': 'Material',
                'LocationOnly': 'LocationOnly',
                'CTD_QTY_Expr': 'CTD_QTY_Expr',
                'Typ Cntner Qty': 'TypicalContainerQty',
                'Typ Plt Qty': 'TypicalPalletQty',
                'Notes': 'Notes',
                'PKGID_Desc': 'PKGID_Desc',
                'TAGQTY': 'TAGQTY',
                'Poss Not Rcvd': 'FLAG_PossiblyNotRecieved',
                'Mvmt Dur Ct': 'FLAG_MovementDuringCount',
                'WICSignore': NOTdbFld_flags[0],
                }
        for col in SprshtcolmnNames:
            if col.value in Sprsht_SSName_TableName_map:
                colkey = Sprsht_SSName_TableName_map[str(col.value)]
                # has this col.value already been mapped?
                if (colkey in SprshtcolmnMap and SprshtcolmnMap[colkey] is not None):
                    # yes, that's a problem
                    self.showUpdateStatus(f"Error: This workbook has a bad header row - More than one column named {col.value}.  See Calvin to fix this.")
                    wb.close()
                    return
                else:
                    SprshtcolmnMap[colkey] = col.column - 1 # type: ignore
                # endif previously mapped
            #endif col.value in SAP_SSName_TableName_map
        #endfor col in SAPcolmnNames

        HeaderGood = all([(reqFld in SprshtcolmnMap) for reqFld in SprshtREQUIREDFLDS])
        if not HeaderGood:
            MissingRequiredFields = [reqFld for reqFld in SprshtREQUIREDFLDS if reqFld not in SprshtcolmnMap]
            self.showUpdateStatus(f"Error: This workbook has a bad header row - missing columns {MissingRequiredFields}.  See Calvin to fix this.")
            wb.close()
            return

        SprshtRowNum=1
        nRowsAdded = 0
        nRowsNoMaterial = 0
        nRowsErrors = 0

        intrval_announce = min(100, int(max(1, ws.max_row // 10)))
        for row in ws.iter_rows(min_row=SprshtRowNum+1, values_only=True):
            SprshtRowNum += 1
            if SprshtRowNum % intrval_announce == 0:
                self.showUpdateStatus(f"Reading Spreadsheet ... record {SprshtRowNum} of {ws.max_row}", SprshtRowNum, ws.max_row)

            ignoreline = ( ( (NOTdbFld_flags[0] in SprshtcolmnMap) and (row[SprshtcolmnMap[NOTdbFld_flags[0]]]) )
                        or (row[SprshtcolmnMap['Material']] is None)
                        )
            if not ignoreline:
                matlnum = self.cleanupfld('Material', row[SprshtcolmnMap['Material']])['cleanval']
                # if no org given, check that Material unique.
                if Sprsht_SSName_TableName_map['org_id'] not in SprshtcolmnMap:
                    spshtorg = None
                else:
                    spshtorg = self.cleanupfld('org_id', row[SprshtcolmnMap['org_id']])['cleanval']
                matlorglist = Repository(get_app_sessionmaker(), MaterialList).get_all(MaterialList.Material==matlnum)
                MatlKount = len(matlorglist)
                MatObj = None
                err_already_handled = False
                if MatlKount == 1:
                    MatObj = matlorglist[0]
                    spshtorg = MatObj.org_id
                if MatlKount > 1:
                    if spshtorg is None:
                        r = UploadSAPResults(
                            errState = 'error',
                            errmsg = f"{matlnum} in multiple org_id's {tuple(matlorglist)}, but no org_id given",
                            rowNum = SprshtRowNum
                            )
                        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
                        nRowsErrors += 1
                        err_already_handled = True
                    else:
                        for mat in matlorglist:
                            if mat.org_id == spshtorg:
                                MatObj = mat
                                break
                    if MatObj is None and not err_already_handled:
                        r = UploadSAPResults(
                            errState = 'error',
                            errmsg = f"{matlnum} in in multiple org_id's {tuple(matlorglist)}, but org_id given ({spshtorg}) is not one of them",
                            rowNum = SprshtRowNum
                            )
                        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
                        nRowsErrors += 1
                        err_already_handled = True
                    #endif spshtorg is None
                #endif MatKount > 1

                if not MatObj:
                    if not err_already_handled:
                        nRowsErrors += 1
                        r = UploadSAPResults(
                            errState = 'error',
                            errmsg = f'either {matlnum} does not exist in MaterialList or incorrect org_id ({str(spshtorg)}) given',
                            rowNum = SprshtRowNum
                            )
                        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
                else:
                    rowErrs = False
                    requiredFields = {reqFld: False for reqFld in SprshtREQUIREDFLDS}
                    requiredFields['Both LocationOnly and CTD_QTY'] = False

                    MatChanged = False
                    SRec = ActualCounts()
                    for fldName, colNum in SprshtcolmnMap.items():
                        if fldName in NOTdbFld_flags: continue
                        # check/correct problematic data types
                        usefld, V = self.cleanupfld(fldName, row[colNum], CountSprshtDateEpoch=CountSprshtDateEpoch).values()
                        if (V is not None):
                            if usefld:
                                if   fldName == 'CountDate':
                                    setattr(SRec, fldName, V)
                                    requiredFields['CountDate'] = True
                                elif fldName == 'Material':
                                    setattr(SRec, fldName, MatObj)
                                    requiredFields['Material'] = True
                                elif fldName == 'Counter':
                                    setattr(SRec, fldName, V)
                                    requiredFields['Counter'] = True
                                elif fldName == 'LOCATION':
                                    setattr(SRec, fldName, V)
                                    requiredFields['LOCATION'] = True
                                elif fldName == 'LocationOnly':
                                    setattr(SRec, fldName, True if V else False)
                                    requiredFields['Both LocationOnly and CTD_QTY'] = True
                                elif fldName == 'CTD_QTY_Expr':
                                    setattr(SRec, fldName, V)
                                    requiredFields['Both LocationOnly and CTD_QTY'] = True
                                elif fldName == 'TypicalContainerQty' \
                                or fldName == 'TypicalPalletQty':
                                    if V == '' or V == None: V = 0
                                    if V != 0 and V != getattr(MatObj,fldName,0):
                                        setattr(MatObj, fldName, V)
                                        MatChanged = True
                                else:
                                    if hasattr(SRec, fldName): setattr(SRec, fldName, V)
                                # endif fldname
                            else:
                                if fldName!='CTD_QTY_Expr':
                                    # we have to suspend judgement on CTD_QTY_Expr until last, because this could be a LocationOnly count
                                    rowErrs = True
                                    r = UploadSAPResults(
                                        errState = 'error',
                                        errmsg = f'{str(V)} is invalid for {fldName}',
                                        rowNum = SprshtRowNum
                                        )
                                    Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
                            #endif usefld
                        #endif (V is not None)
                    # for each column

                    # now we determine if one of LocationOnly or CTD_QTY was given
                    if not requiredFields['Both LocationOnly and CTD_QTY']:
                        fldName = 'CTD_QTY_Expr'
                        V = row[SprshtcolmnMap[fldName]]
                        rowErrs = True
                        r = UploadSAPResults(
                            errState = 'error',
                            errmsg = f'record is not marked LocationOnly and {str(V)} is invalid for {fldName}',
                            rowNum = SprshtRowNum
                            )
                        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)

                    # are all required fields present?
                    AllRequiredPresent = True
                    for keyname, Prsnt in requiredFields.items():
                        AllRequiredPresent = AllRequiredPresent and Prsnt
                        if not Prsnt:
                            rowErrs = True
                            r = UploadSAPResults(
                                errState = 'error',
                                errmsg = f'{keyname} missing',
                                rowNum = SprshtRowNum
                                )
                            Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
                    # endfor requiredFields

                    if not rowErrs:
                        Repository(get_app_sessionmaker(), ActualCounts).add(SRec)
                        if MatChanged: Repository(get_app_sessionmaker(), MaterialList).update(MatObj)
                        resultString = str(SRec)
                        resultString += ' / LOCATION ONLY'  if SRec.LocationOnly else f' / Qty= {SRec.CTD_QTY_Expr}'
                        resultString += ' (Typ Cont Qty/Typ Plt Qty also changed)' if MatChanged else ''
                        r = UploadSAPResults(
                            errState = 'success',
                            errmsg = resultString,
                            rowNum = SprshtRowNum
                            )
                        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
                        nRowsAdded += 1
                    else:
                        nRowsErrors += 1
                    #endif not rowErrs
                # endif MatObj/not MatObj
            else:
                nRowsNoMaterial += 1
            #endif not ignoreline
        # endfor row in ws.iter_rows

        r = UploadSAPResults(
            errState = 'nRowsTotal',
            errmsg = '',
            rowNum = SprshtRowNum
            )
        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
        r = UploadSAPResults(
            errState = 'nRowsAdded',
            errmsg = '',
            rowNum = nRowsAdded
            )
        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
        r = UploadSAPResults(
            errState = 'nRowsErrors',
            errmsg = '',
            rowNum = nRowsErrors
            )
        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)
        r = UploadSAPResults(
            errState = 'nRowsIgnored',
            errmsg = '',
            rowNum = nRowsNoMaterial
            )
        Repository(get_app_sessionmaker(), UploadSAPResults).add(r)

        # close and kill temp files
        wb.close()

        # self.done_UpActCountSprsheet_01ReadSheet()    # caller will do this
    def done_UpActCountSprsheet_01ReadSheet(self):
        self.showUpdateStatus("Finished Reading Count Entry Spreadsheet.")
        self.proc_UpActCountSprsheet_99_FinalProc()
        #endif stateocde != 'fatalerr'

    def proc_UpActCountSprsheet_99_FinalProc(self) -> None:
        self.showUpdateStatus("Finished Processing Count Entry Spreadsheet...")

    def proc_UpActCountSprsheet_99_Cleanup(self) -> None:
        # delete the temporary table
        Repository(get_app_sessionmaker(), UploadSAPResults).removewhere(True)
# UploadActCountSprsht
    def end_of_class(self):
        pass
    
class ShowUploadedCountEntries(QWidget):
    """A form to show the Update Material List from SAP MM60 or ZMSQV001 Spreadsheet form."""
    _formname = "Upload Count Spreadsheet - Results"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.setWindowTitle(self._formname)
        myLayout = QVBoxLayout(self)
        lblResultsTitle = QLabel(self._formname)
        myLayout.addWidget(lblResultsTitle)

        wdgtMainArea = QWidget()
        layoutMainArea = QVBoxLayout(wdgtMainArea)
        wdgtScrollArea = QScrollArea()
        wdgtScrollArea.setWidgetResizable(True)
        wdgtScrollArea.setWidget(wdgtMainArea)
        myLayout.addWidget(wdgtScrollArea)

        statusVal = Repository(get_app_sessionmaker(), UploadSAPResults).get_all(
            UploadSAPResults.errState == 'nRowsTotal'
        )
        nRowsRead = statusVal[0].rowNum - 1 if statusVal else 0     # -1 because header doesn't count
        statusVal = Repository(get_app_sessionmaker(), UploadSAPResults).get_all(
            UploadSAPResults.errState == 'nRowsAdded'
        )
        nRowsAdded = statusVal[0].rowNum if statusVal else 0
        statusVal = Repository(get_app_sessionmaker(), UploadSAPResults).get_all(
            UploadSAPResults.errState == 'nRowsErrors'
        )
        nRowsErrors = statusVal[0].rowNum if statusVal else 0
        statusVal = Repository(get_app_sessionmaker(), UploadSAPResults).get_all(
            UploadSAPResults.errState == 'nRowsIgnored'
        )
        nRowsNoMaterial = statusVal[0].rowNum if statusVal else 0

        UplResults = Repository(get_app_sessionmaker(), UploadSAPResults).get_all(
            UploadSAPResults.errState.notin_(['nRowsAdded','nRowsTotal','nRowsErrors','nRowsIgnored'])
        )
        lblSummary = QLabel(f"Upload Summary: {nRowsRead} rows read, {nRowsAdded} rows added, {nRowsErrors} rows with errors, {nRowsNoMaterial} rows ignored (no material).")
        layoutMainArea.addWidget(lblSummary)

        for res in UplResults:
            rstr = f"{res.errState}: {res.errmsg}" if 'error' in res.errState else f"Count Record {res.errmsg} added"
            lblrslt = QLabel(rstr)
            layoutMainArea.addWidget(lblrslt)
        # endif UplResults
    # __init__
# ShowUpdateMatlListfromSAPForm
    def end_of_class(self):
        pass


class UploadSAPSOHSprsht(cSRFSingleRecordForm):
    _ORMmodel = SAP_SOHRecs
    _formname = "Upload SAP MB52 Spreadsheet"
    _ssnmaker = get_app_sessionmaker()

    LastSAPUpload = None
    uplDate = None
    uplDateWarning = None
    btnChooseFile = None
    wdgtUpdtStatusText = None
    wdgtUpdtStatusProgBar = None


    # __init__ inherited from cSimpleRecordForm
    def __init__(self, *args, **kwargs):
        self.uploadresults: dict[str, Any] = {}

        super().__init__(*args, **kwargs)
    # __init__

    def defineFields(self) -> List[cQFormFieldDef] | None:
        # no fields to edit, everything manually handled
        return []


    ###########################################################
    ############## Override cSimpleRecordForm UI build methods
    ###########################################################

    @Slot()
    def _build_fields(self) -> None:
    # def _placeFields(self, layoutFormPages: QTabWidget, layoutFormFixedTop: QGridLayout | None, layoutFormFixedBottom: QGridLayout | None, lookupsAllowed: bool = True) -> None:
        # no fields to place, everything manually handled
        mainFormPage = self.FormPage(0)
        assert isinstance(mainFormPage, QGridLayout)

        layoutFormFixedTop = self._layouts.fixed_top

        # one day, Ill have the db do this ...
        SAPDates = { dd.uploaded_at for dd in Repository(get_app_sessionmaker(), SAP_SOHRecs).get_all() }
        self.LastSAPUpload = None if not SAPDates else max(SAPDates)
        anncestrLastDate = "No Previous SAP SOH Upload" if self.LastSAPUpload is None else f"Last SAP MB52 Upload was on: {self.LastSAPUpload.strftime('%Y-%m-%d')}"
        mainFormPage.addWidget(QLabel(anncestrLastDate), 0, 0)

        todaysdate = datetime.now().date()
        self.uplDate = QDateEdit(date=QDate(todaysdate.year, todaysdate.month, todaysdate.day))
        self.uplDate.userDateChanged.connect(self.uplDateChanged)
        uplDateTitle = QLabel("Upload Date (for SAP SOH records):")
        self.uplDateWarning = QLabel(" ")
        self.uplDateWarning.setStyleSheet("color: red; font-weight: bold;")
        mainFormPage.addWidget(uplDateTitle, 1, 0)
        mainFormPage.addWidget(self.uplDate, 1, 1)
        mainFormPage.addWidget(self.uplDateWarning, 1, 2)

        chooseFileWidget = QWidget()
        chooseFileLayout = QGridLayout(chooseFileWidget)
        lblChooseFileLabel = QLabel("Choose or Drop SAP MB52 Spreadsheet File:")
        self.btnChooseFile = cFileSelectWidget(btnText="Choose Spreadsheet File")
        self.btnChooseFile.fileChosen.connect(self.FileChosen)
        chooseFileLayout.addWidget(lblChooseFileLabel, 0, 0)
        chooseFileLayout.addWidget(self.btnChooseFile, 1, 0)
        mainFormPage.addWidget(chooseFileWidget, 2, 0)

        # the status area
        self.wdgtUpdtStatusText = QLabel("")
        self.wdgtUpdtStatusProgBar = QProgressBar()
        assert layoutFormFixedTop is not None, "layoutFormFixedTop is required"
        layoutFormFixedTop.addWidget(self.wdgtUpdtStatusText, 0, 0)
        layoutFormFixedTop.addWidget(self.wdgtUpdtStatusProgBar, 1, 0)
    # _placeFields


    def _addActionButtons(self, ActionButtons) -> None:     # pylint:disable=signature-differs
        layoutButtons = self._layouts.buttons

        btnUploadFile = QPushButton(icon('fa5s.file-upload'), "Upload SAP SOH")
        btnUploadFile.clicked.connect(self.uploadFile)
        btnClose = QPushButton(icon('mdi.close-octagon'), "Close Form")
        btnClose.clicked.connect(self.closeForm)

        if layoutButtons is not None:
            layoutButtons.addWidget(btnUploadFile, alignment=Qt.AlignmentFlag.AlignLeft)
            layoutButtons.addStretch(1)
            layoutButtons.addWidget(btnClose, alignment=Qt.AlignmentFlag.AlignRight)
        #endif layoutButtons
    # def _handleActionButton(self, action: str) -> None:
    #     return super()._handleActionButton(action)
    # _add/handleActionButtons

    def initialdisplay(self):
        self.showNewRecordFlag()
        self.showUpdateStatus("")
    # initialdisplay

    ###########################################################
    ############## UI update methods
    ###########################################################

    def showUpdateStatus(self, statusText: str, progressValue: int = 0, progressMax: int = -1) -> None:

        assert self.wdgtUpdtStatusText is not None, "wdgtUpdtStatusText is not defined"
        assert self.wdgtUpdtStatusProgBar is not None, "wdgtUpdtStatusProgBar is not defined"

        self.wdgtUpdtStatusText.setText(statusText)

        if progressMax < 0:
            self.wdgtUpdtStatusProgBar.setVisible(False)
        else:
            self.wdgtUpdtStatusProgBar.setVisible(True)
            self.wdgtUpdtStatusProgBar.setMaximum(progressMax)
            self.wdgtUpdtStatusProgBar.setValue(progressValue)
        # endif progressMax
    # showUpdateStatus


    ###########################################################
    ############## button responders
    ###########################################################

    @Slot()
    def FileChosen(self):
        # enable the Upload button
        pass

    @Slot()
    def uplDateChanged(self, dateslctd: QDate):
        assert self.uplDateWarning is not None, "uplDateWarning widget is not defined"

        # Check if SAP SOH records already exist for this date
        UplDate = datetime(dateslctd.year(), dateslctd.month(), dateslctd.day()).date()
        self.uplDateWarning.setText(" ")    # assume no warning
        existingRecs = Repository(get_app_sessionmaker(), SAP_SOHRecs).get_all(SAP_SOHRecs.uploaded_at==UplDate)
        if len(existingRecs) > 0:
            self.uplDateWarning.setText(f"(an existing SAP upload exists for {UplDate.strftime('%Y-%m-%d')}. It will be overwritten!!)")


    @Slot()
    def uploadFile(self):
        # disable most of the form while processing

        self.proc_UpSAPSprsheet_00InitUpld()

        USSName = self.proc_UpSAPSprsheet_00CopySpreadsheet()
        self.proc_UpSAPSprsheet_01ReadSheet(USSName)
        self.done_UpSAPSprsheet_01ReadSheet()      # this triggers most of the rest of the processing chain

        # present results to user
        childScreen = ShowUploadedSAPResults(uploadresults=self.uploadresults)
        childScreen.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        childScreen.show()

        self.proc_UpSAPSprsheet_99_Cleanup()
        self.closeForm()


    @Slot()
    def closeForm(self):
        # TODO: Implement close form logic
        self.close()


    ###########################################################
    ############## "record" status methods
    ###########################################################

    def isNewRecord(self) -> bool:
        return False


    ###########################################################
    ############## other cSimpleRecordForm overrides
    ###########################################################

    def changeInternalVarField(self, wdgt, intVarField: str, wdgt_value: Any) -> None:      # pylint:disable=unused-argument
        return


    ###########################################################
    ############## Read / process spreadsheet methods
    ###########################################################

    def proc_UpSAPSprsheet_00InitUpld(self) -> None:
        self.showUpdateStatus("Initializing Upload of SAP MB52 Spreadsheet...")

    def proc_UpSAPSprsheet_00CopySpreadsheet(self) -> str:
        assert self.btnChooseFile is not None, "btnChooseFile widget is not defined"
        return self.btnChooseFile.getFileChosen()

    def proc_UpSAPSprsheet_01ReadSheet(self, fName: str) -> None:
        self.showUpdateStatus("Reading SAP MB52 Spreadsheet...")

        _SStName_Material = 'Material'
        _db_Name_Material = 'MaterialPartNum'
        _SStName_Plant = 'Plant'
        _db_Name_Plant = 'Plant'

        def SAPFldDescMap() -> Dict[str,cExcelFile.SprdsheetFldDescriptor]:
            Sprsht_SSName_TableName_map = {
                    # Material+org will translate to a Material_id
                    _SStName_Material: {'ModelFldName': _db_Name_Material},
                    _SStName_Plant: {'ModelFldName': _db_Name_Plant},
                    'OrgID': {'ModelFldName': 'org_id', 'CalculatedFld': True, 'CleanProc': SAPCalcFldProc, 'AllowedTypes': (int, )},
                    'MatlID': {'ModelFldName': 'Material_id', 'CalculatedFld': True, 'CleanProc': SAPCalcFldProc, 'AllowedTypes': (int, )},
                    'UpldAt': {'ModelFldName': 'uploaded_at', 'CalculatedFld': True, 'CleanProc': SAPCalcFldProc, 'AllowedTypes': (date, )},
                    'Material description': {'ModelFldName': 'Description'},
                    'Material type': {'ModelFldName': 'MaterialType'},
                    'Storage location': {'ModelFldName': 'StorageLocation'},
                    'Base Unit of Measure': {'ModelFldName': 'BaseUnitofMeasure'},
                    'Unrestricted': {'ModelFldName': 'Amount', 'CleanProc': SAPFldCleanProc, 'AllowedTypes': (float, int, )},
                    'Currency': {'ModelFldName': 'Currency'},
                    'Value Unrestricted': {'ModelFldName': 'ValueUnrestricted', 'CleanProc': SAPFldCleanProc, 'AllowedTypes': (float, int, )},
                    'Special Stock': {'ModelFldName': 'SpecialStock'},
                    'Blocked':{'ModelFldName': 'Blocked', 'CleanProc': SAPFldCleanProc, 'AllowedTypes': (float, int, )},
                    'Value BlockedStock':{'ModelFldName': 'ValueBlocked', 'CleanProc': SAPFldCleanProc, 'AllowedTypes': (float, int, )},
                    'Vendor':{'ModelFldName': 'Vendor'},
                    'Batch': {'ModelFldName': 'Batch'},
                    }
            retDict = {ssName: cExcelFile.SprdsheetFldDescriptor(**fldDescArgs)
                for ssName, fldDescArgs in Sprsht_SSName_TableName_map.items()}
            return retDict
        # SAPFldDescMap

        def SAPFldCleanProc(spshtFldNm: str, dbFldNm: str, fldVal: Any, rowdict) -> Any:        # pylint:disable=unused-argument
            if fldVal is None:
                return None
            if spshtFldNm in ['Unrestricted', 'Value Unrestricted', 'Value BlockedStock', 'Blocked']:
                try:
                    return float(fldVal)
                except ValueError:
                    return 0.0
            # endif numeric fields
            return fldVal
        # SAPFldCleanProc

        def SAPCalcFldProc(spshtFldNm:str, dbFldNm:str, val:Any, rowdict) -> Any:    # pylint:disable=unused-argument
            if spshtFldNm == 'OrgID':
                plant_val = rowdict[_db_Name_Plant]
                org_list = Repository(get_app_sessionmaker(), SAPPlants_org).get_all(SAPPlants_org.SAPPlant==plant_val)
                if len(org_list) < 1:
                    return None
                org_id = org_list[0].org_id
                return org_id
            elif spshtFldNm == 'MatlID':
                plant_val = rowdict[_db_Name_Plant]
                _org = Repository(get_app_sessionmaker(), SAPPlants_org).get_all(SAPPlants_org.SAPPlant==plant_val)[0].org_id
                MatlNum = rowdict[_db_Name_Material]
                try:
                    whereclause = and_(MaterialList.org_id==_org, MaterialList.Material==MatlNum)
                    MatlRec = Repository(get_app_sessionmaker(), MaterialList).get_all(whereclause)[0]
                    return MatlRec.id
                except (TypeError, IndexError, KeyError):
                    return None
            elif spshtFldNm == 'UpldAt':
                assert self.uplDate is not None, "uplDate is not defined"
                return self.uplDate.date().toPython()
            # endif calculated fields
            return None
        # SAPCalcFldProc

        def UplSAPProgressAnnounceCallback(currentRowNum, totalRows):
            self.showUpdateStatus(f"Reading Spreadsheet ... record {currentRowNum} of {totalRows}", currentRowNum, totalRows)

        # if SAP SOH records exist for this date, kill them; only one set of SAP SOH records per day
        # (this was signed off on by user before coming here)
        assert self.uplDate is not None, "uplDate is not defined"
        UplDate = self.uplDate.date().toPython()
        Repository(get_app_sessionmaker(), SAP_SOHRecs).removewhere(SAP_SOHRecs.uploaded_at==UplDate)

        # wb = load_workbook(filename=fName, read_only=True)
        wb = cExcelFile.load_from_file(filename=fName, read_only=True)
        assert wb is not None, "Failed to load spreadsheet file"
        # CountSprshtDateEpoch = wb.epoch

        wb.save_to_SQLAlchemyModel(
            ssnmaker=self._ssnmaker,
            TargetModel=SAP_SOHRecs,
            WksheetName=None,   # default to active sheet
            SprdsheetFlds=SAPFldDescMap(),
            required_columns=[_db_Name_Material, _db_Name_Plant, ],
            progress_interval=100,
            progress_callback=UplSAPProgressAnnounceCallback,
            )

        self.uploadresults['nRowsTotal'] = wb.num_rows
        self.uploadresults['nRowsAdded'] = wb.nRows
        self.uploadresults['uploadDate'] = UplDate

        # close and kill temp files
        wb.close()

        # self.done_UpSAPSprsheet_01ReadSheet()    # caller will do this
    def done_UpSAPSprsheet_01ReadSheet(self):
        self.showUpdateStatus("Finished Reading Count Entry Spreadsheet.")
        self.proc_UpSAPSprsheet_99_FinalProc()

    def proc_UpSAPSprsheet_99_FinalProc(self) -> None:
        self.showUpdateStatus("Finished Processing Count Entry Spreadsheet...")

    def proc_UpSAPSprsheet_99_Cleanup(self) -> None:
        # nothing to clean up here
        return
# UploadSAPSOHSprsht
    def end_of_class(self):
        pass

class ShowUploadedSAPResults(QWidget):
    """A form to show the Upload SAP SOH Spreadsheet results."""
    _formname = "Upload SAP SOH Spreadsheet - Results"

    def __init__(self, *args, **kwargs):
        uploadresults = kwargs.pop('uploadresults', {})
        super().__init__(*args, **kwargs)

        self.setWindowTitle(self._formname)
        myLayout = QVBoxLayout(self)
        lblResultsTitle = QLabel(self._formname)
        myLayout.addWidget(lblResultsTitle)

        wdgtMainArea = QWidget()
        layoutMainArea = QVBoxLayout(wdgtMainArea)
        wdgtScrollArea = QScrollArea()
        wdgtScrollArea.setWidgetResizable(True)
        wdgtScrollArea.setWidget(wdgtMainArea)
        myLayout.addWidget(wdgtScrollArea)

        # nRowsRead = uploadresults.get('nRowsTotal', 0)
        nRowsAdded = uploadresults.get('nRowsAdded', 0)
        upldDate = uploadresults.get('uploadDate', None)

        lblSummary = QLabel(
            f"<h4>{ nRowsAdded } SAP SOH (MB52) spreadsheet records successfully uploaded with date { upldDate }!</h4>"
            )
        layoutMainArea.addWidget(lblSummary)
    # __init__
# ShowUploadedSAPResults
    def end_of_class(self):
        pass

